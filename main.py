"""
@Description :
@File        : main.py
@Project     : ClacLine
@Time        : 2022/4/3 15:30
@Author      : hanhan
@Software    : PyCharm
"""

import os

import openpyxl

from Configs.Config import *
from Tools import *


def main():
    # 计算需要统计的文件夹的数量
    startPos = 0
    endCurvePos = len([lists for lists in os.listdir(curvePath) if os.path.isdir(os.path.join(curvePath, lists))])
    endPolylinePos = len(
        [lists for lists in os.listdir(polylinePath) if os.path.isdir(os.path.join(polylinePath, lists))])
    # endCurvePos = 8 + 1
    # endPolylinePos = -1 + 1

    # 处理曲线图
    for pos in range(startPos, endCurvePos):
        # 各个地址
        imgPath = curvePath + '/' + str(pos) + '/' + drawFileName
        drawWhitePath = curvePath + '/' + str(pos) + '/' + drawWhiteFileName
        borderPath = curvePath + '/' + str(pos) + '/' + borderFileName
        drawLinePath = curvePath + '/' + str(pos) + '/' + drawLineFileName
        dbPath = curvePath + '/' + str(pos) + '/' + dbFileName
        drawMarkPath = curvePath + '/' + str(pos) + '/' + drawMarkFileName
        ansPath = curvePath + '/' + str(pos) + '/' + ansFileName
        drawMarkPath2 = MarkImgPath2 + '/' + curvePath + '/' + str(pos) + '.png'
        drawTwoPath = curvePath + '/' + str(pos) + '/' + drawTwoFileName

        print("正在处理图片：" + imgPath)

        # img = cv2.imread(imgPath);
        # b, g, r = cv2.split(img);
        # cv2.imshow('Blue', b);
        # cv2.imshow('Green', g);
        # cv2.imshow('Red', r);
        # cv2.waitKey(0)

        # 白化
        TurnWhite(imgPath, drawWhitePath)

        # 获取边框
        GetBorder(drawWhitePath, borderPath)

        # 二选一
        # 去除边框颜色
        ClearBorder(borderPath, drawWhitePath, drawLinePath)
        # 二值化
        # TurnTwo(drawWhitePath, drawTwoPath)

        # 获取五个点得结果
        GetAns(drawLinePath, borderPath, dbPath, drawMarkPath, ansPath, drawMarkPath2)
        # GetAns(drawTwoPath, borderPath, dbPath, drawMarkPath, ansPath, drawMarkPath2)

    # 处理折线图
    for pos in range(startPos, endPolylinePos):
        # 各个地址
        imgPath = polylinePath + '/' + str(pos) + '/' + drawFileName
        drawWhitePath = polylinePath + '/' + str(pos) + '/' + drawWhiteFileName
        borderPath = polylinePath + '/' + str(pos) + '/' + borderFileName
        drawLinePath = polylinePath + '/' + str(pos) + '/' + drawLineFileName
        dbPath = polylinePath + '/' + str(pos) + '/' + dbFileName
        drawMarkPath = polylinePath + '/' + str(pos) + '/' + drawMarkFileName
        ansPath = polylinePath + '/' + str(pos) + '/' + ansFileName
        drawMarkPath2 = MarkImgPath2 + '/' + polylinePath + '/' + str(pos)
        drawTwoPath = polylinePath + '/' + str(pos) + '/' + drawTwoFileName

        # 白化
        TurnWhite(imgPath, drawWhitePath)

        # 获取边框
        GetBorder(drawWhitePath, borderPath)

        # 去除边框颜色
        ClearBorder(borderPath, drawWhitePath, drawLinePath)
        # 二值化
        # TurnTwo(drawWhitePath, drawTwoPath)

        # 获取五个点得结果
        GetAns(drawLinePath, borderPath, dbPath, drawMarkPath, ansPath, drawMarkPath2)
        # GetAns(drawTwoPath, borderPath, dbPath, drawMarkPath, ansPath, drawMarkPath2)


def binary():
    lenCurve = len([lists for lists in os.listdir(curvePath) if os.path.isdir(os.path.join(curvePath, lists))])
    lenPolyline = len([lists for lists in os.listdir(polylinePath) if os.path.isdir(os.path.join(polylinePath, lists))])
    # print(lenCurve)
    for pos in range(lenCurve):
        imgPath = curvePath + '/' + str(pos) + '/' + drawFileName  # 原图片地址
        cutImgPath = cutBorderPath + '/Curve/' + str(pos) + '.png'
        binaryPath = binImgPath + '/Curve/' + str(pos) + '.png'
        blackBinaryPath = blackPath + '/Curve/' + str(pos) + '.png'
        delpath = delNetPath + '/Curve/' + str(pos) + '.png'
        dbpath = curvePath + '/' + str(pos) + '/' + dbFileName
        ansPath = ansBinPath + '/Curve/' + str(pos) + '.txt'
        markpath = markBinPath + '/Curve/' + str(pos) + '.png'

        print('正在处理' + imgPath)
        CutBorder(imgPath, cutImgPath)
        TurnBinary(cutImgPath, binaryPath)
        TurnBlackBackground(binaryPath, blackBinaryPath)
        DeleteLine(blackBinaryPath, delpath)
        GetBinaryAns(delpath, markpath, dbpath, ansPath)

    for pos in range(lenPolyline):
        imgPath = polylinePath + '/' + str(pos) + '/' + drawFileName
        cutImgPath = cutBorderPath + '/Polyline/' + str(pos) + '.png'
        binaryPath = binImgPath + '/Polyline/' + str(pos) + ".png"
        blackBinaryPath = blackPath + '/Polyline/' + str(pos) + '.png'
        delpath = delNetPath + '/Polyline/' + str(pos) + '.png'
        dbpath = polylinePath + '/' + str(pos) + '/' + dbFileName
        ansPath = ansBinPath + '/Polyline/' + str(pos) + '.txt'
        markpath = markBinPath + '/Polyline/' + str(pos) + '.png'

        print('正在处理' + imgPath)
        CutBorder(imgPath, cutImgPath)
        TurnBinary(cutImgPath, binaryPath)
        TurnBlackBackground(binaryPath, blackBinaryPath)
        DeleteLine(blackBinaryPath, delpath)
        GetBinaryAns(delpath, markpath, dbpath, ansPath)


def test():
    # DeleteLine('img.png', 'ans.png')
    CutBorder('draw.png', 'ans.png')
    # TurnBlackBackground(binaryPath, blackBinaryPath)
    # TurnBinary('ans.png', 'ans.png')
    # GetBinaryAns('Ans/DelNet/Curve/10.png', 'ans.png', 'TrainData/Curve/10/db.txt', 'ans.txt')


def PrintAns():
    wb = openpyxl.Workbook()
    sheet1 = wb.create_sheet('Sheet1')
    for row in range(500):
        curve = open('Ans/Ans/Curve/' + str(row) + '.txt', 'r', encoding='utf8')
        ls = list(map(float, curve.read().split()))
        if len(ls) < 5:
            maxSize = int(open('Data/Curve/' + str(row) + '/db.txt', 'r', encoding='utf8').read().split()[0]) // 2
            ls = [maxSize, maxSize, maxSize, maxSize, maxSize]
        # print(ls)
        for col, item in enumerate(ls):
            sheet1.cell(row = row + 1, column = col + 1).value = item
    sheet2 = wb.create_sheet('Sheet2')
    for row in range(500):
        polyline = open('Ans/Ans/Polyline/' + str(row) + '.txt', 'r', encoding='utf8')
        ls = list(map(float, polyline.read().split()))
        if len(ls) < 5:
            maxSize = int(open('Data/Polyline/' + str(row) + '/db.txt', 'r', encoding='utf8').read().split()[0]) // 2
            ls = [maxSize, maxSize, maxSize, maxSize, maxSize]
        # print(ls)
        for col, item in enumerate(ls):
            sheet2.cell(row = row + 1, column = col + 1).value = item
    wb.save('ans.xlsx')



if __name__ == '__main__':
    # main()
    # binary()
    # test()
    PrintAns()

